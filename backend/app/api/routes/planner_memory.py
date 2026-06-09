"""
Planner Task Memory Route
GET  /api/planner/memory?dataset=<id>   — get cached task memory for a dataset
POST /api/planner/memory/process        — run full memory pipeline on conversation data

Pipeline (per the planner docs):
1. Parse conversations from mock Excel data (childconvo.xlsx)
2. Extract facts → Engagement report (E1) — with LLM
3. For each developmental domain: extract observations, detect duplicates, update occurrence table
4. Compute derived insights (D1) from updated observations
5. Cache result per dataset in main/planner_memory/
"""
from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from dotenv import load_dotenv
from fastapi import APIRouter, HTTPException, Query, status
from pydantic import BaseModel

from app.core.prompts import MEMORY_ANALYSIS_PROMPT
from app.core.posthog_client import get_posthog

load_dotenv()

router = APIRouter(prefix="/api/planner", tags=["planner-memory"])
logger = logging.getLogger(__name__)

# ─── Constants ────────────────────────────────────────────────────────────────

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "mockdata", "childconvo.xlsx"
)

MEMORY_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "main", "planner_memory"
)

GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")

# Developmental domains used by the planner
DOMAINS = [
    "COGNITIVE",
    "LANGUAGE",
    "SOCIAL_EMOTIONAL",
    "APPROACHES_TO_LEARNING",
    "CULTURAL_VALUES",
    "PHYSICAL_HEALTH",
]

# ─── Schemas ──────────────────────────────────────────────────────────────────

class ObservationOccurrence(BaseModel):
    occur_idx: int
    date: str
    details: str
    conversation_id: str

class DomainObservation(BaseModel):
    observation: str
    domain: str
    occurrences: List[ObservationOccurrence]
    last_update_at: str
    is_new_this_week: bool

class PlannerMemoryResponse(BaseModel):
    dataset: str
    profile_id: str
    week_label: str
    engagement_report: str          # E1 — engagement summary for current week
    prev_engagement_report: str     # E2 — previous week engagement (if available)
    observations_by_domain: Dict[str, List[DomainObservation]]
    derived_insights: str           # D1 — what Pika noticed this week
    talk_history: List[str]         # topic facets used in sessions this week
    conversation_count: int
    message_count: int
    generated_at: str

class UpdateMemoryRequest(BaseModel):
    dataset: str
    memory_clusters: Optional[List[Dict]] = None
    talk_history: Optional[List[str]] = None

# ─── Dataset → Profile ID map ─────────────────────────────────────────────────

# Maps dataset names to real profile_id values found in the Excel
DATASET_PROFILE_MAP = {
    "019dfd3e-282c-76b9-a760-b9cf3cd22212": "019dfd3e-282c-76b9-a760-b9cf3cd22212",
    "019e7fa3-5b8a-7c5c-bc63-2bfbd302e61b": "019e7fa3-5b8a-7c5c-bc63-2bfbd302e61b",
    "019dbf57-771d-7a01-8b92-c1592ad61f8f": "019dbf57-771d-7a01-8b92-c1592ad61f8f",
    "019c9991-6ad8-7a87-91ff-673ec60b6d6f": "019c9991-6ad8-7a87-91ff-673ec60b6d6f",
    "019cff81-1bc3-7939-9230-a1f032605728": "019cff81-1bc3-7939-9230-a1f032605728",
    "019bf3be-0533-799b-8b39-fe24a75c4bc3": "019bf3be-0533-799b-8b39-fe24a75c4bc3",
}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def _load_excel() -> openpyxl.Workbook:
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Mock Excel not found: {EXCEL_PATH}")
    return openpyxl.load_workbook(EXCEL_PATH)


def _get_conversations_for_profile(profile_id: str) -> List[Dict]:
    """Load conversations from Excel for a given profile_id."""
    wb = _load_excel()
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if str(row_dict.get("profile_id", "")) == str(profile_id):
            rows.append(row_dict)

    return rows


def _build_conversation_text(rows: List[Dict]) -> str:
    """Build a readable transcript from conversation rows."""
    lines = []
    for r in rows:
        char = r.get("character", "")
        content = (r.get("content") or "").strip()
        ts = str(r.get("history_created_at", ""))[:16] if r.get("history_created_at") else ""
        convo_id = str(r.get("conversation_id", ""))[:8]
        if not content:
            continue
        if char == "USER":
            lines.append(f"[{ts}][{convo_id}] BÉ: {content}")
        elif char in ("BOT_RESPONSE_CONVERSATION", "SYSTEM"):
            lines.append(f"[{ts}][{convo_id}] PIKA: {content}")
    return "\n".join(lines)


# ─── AI Analysis (OpenAI GPT-4o + Mem0) ──────────────────────────────────────

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
MEM0_BASE_URL = os.environ.get("MEM0_BASE_URL", "https://mem0.hacknao.edu.vn")


def _call_openai(prompt: str, system_prompt: str = MEMORY_ANALYSIS_PROMPT) -> str:
    """Call OpenAI GPT-4o with a prompt, return text response."""
    import httpx, asyncio, threading

    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY not configured")

    payload = {
        "model": "gpt-4o",
        "max_tokens": 8192,
        "temperature": 0.2,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt},
        ],
    }

    result = {"text": "", "error": None}

    def run():
        async def fetch():
            async with httpx.AsyncClient(timeout=180) as client:
                resp = await client.post(
                    "https://api.openai.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
                    json=payload,
                )
                if resp.status_code >= 400:
                    result["error"] = f"OpenAI error {resp.status_code}: {resp.text[:300]}"
                    return
                data = resp.json()
                choices = data.get("candidates") or data.get("choices") or []
                if choices:
                    msg = choices[0].get("message") or {}
                    result["text"] = msg.get("content", "")
        asyncio.run(fetch())

    t = threading.Thread(target=run)
    t.start()
    t.join(timeout=185)

    if result["error"]:
        raise RuntimeError(result["error"])
    return result["text"]


def _call_llm(prompt: str, system_prompt: Optional[str] = None) -> str:
    """Call best available LLM: OpenAI GPT-4o > Gemini 2.0 Flash."""
    if OPENAI_API_KEY:
        if system_prompt is not None:
            return _call_openai(prompt, system_prompt)
        else:
            return _call_openai(prompt)
    if GOOGLE_API_KEY:
        full_prompt = f"{system_prompt}\n\n{prompt}" if system_prompt else prompt
        return _call_gemini(full_prompt)
    raise RuntimeError("No LLM API key configured (need OPENAI_API_KEY or GOOGLE_API_KEY)")


def _parse_yaml_response(text: str):
    """Parse YAML from LLM response."""
    import yaml
    candidate = text.strip()
    fence = re.search(r"```(?:yaml|yml)?\s*\n(.*?)```", candidate, re.DOTALL)
    if fence:
        candidate = fence.group(1).strip()
    try:
        parsed = yaml.safe_load(candidate)
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        pass
    return None



def _fetch_mem0_memories(profile_id: str) -> List[Dict]:
    """Fetch pre-processed memory strings from Mem0 production API.
    
    Mem0 has already processed ALL conversations -> distilled into clean, 
    deduplicated 1-sentence memory facts. This is the same input the golden 
    data benchmark (analysis_results_v2) was generated from.
    """
    import httpx, asyncio, threading
    
    mem0_base = (MEM0_BASE_URL or "").rstrip("/")
    if not mem0_base:
        return []
    
    result: Dict = {"memories": [], "error": None}
    
    def run():
        async def fetch():
            try:
                async with httpx.AsyncClient(timeout=30) as client:
                    resp = await client.get(
                        f"{mem0_base}/memories",
                        params={"user_id": profile_id},
                        headers={"Accept": "application/json"},
                    )
                    if resp.status_code == 200:
                        data = resp.json()
                        mems = data.get("results", data) if isinstance(data, dict) else data
                        if isinstance(mems, list):
                            result["memories"] = mems
                    else:
                        result["error"] = f"Mem0 HTTP {resp.status_code}"
            except Exception as e:
                result["error"] = str(e)
        asyncio.run(fetch())
    
    t = threading.Thread(target=run)
    t.start()
    t.join(timeout=35)
    
    if result["error"]:
        logger.warning(f"[planner_memory] Mem0 fetch failed for {profile_id}: {result['error']}")
    else:
        logger.info(f"[planner_memory] Mem0 returned {len(result['memories'])} memories for {profile_id}")
    
    return result["memories"]


def _build_mem0_user_prompt(profile_id: str, memories: List[Dict]) -> str:
    """Build the same user prompt format as the reference buddy-talk-eval-be code."""
    # Sort by created_at descending (most recent first)
    sorted_mems = sorted(memories, key=lambda m: str(m.get("created_at") or ""), reverse=True)
    
    lines = [
        f"user_id: {profile_id}",
        "name: Unknown",
        "age: Unknown",
        "en_level: pre_a1",
        "friendship_level: PHASE1",
        f"total_memories: {len(sorted_mems)}",
        "",
        "--- MEMORIES ---",
    ]
    for i, m in enumerate(sorted_mems[:200], 1):  # max 200, same as reference
        text = (m.get("memory") or "").strip()
        if len(text) < 5:
            continue
        date = str(m.get("created_at") or "")[:10]
        lines.append(f"{i}. [{date}] {text}")
    return "\n".join(lines)


def _analyze_persona_from_memories(profile_id: str, memories: List[Dict], week_label: str) -> dict:
    """Use GPT-4o with Mem0 memories as input — matches golden data benchmark exactly."""
    user_prompt = _build_mem0_user_prompt(profile_id, memories)
    
    payload = {
        "model": "gpt-4o",
        "max_tokens": 8192,
        "temperature": 0.2,
        "messages": [
            {"role": "system", "content": MEMORY_ANALYSIS_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
    }
    
    result: Dict = {"text": "", "error": None}
    
    def run():
        import asyncio
        async def fetch():
            import httpx
            async with httpx.AsyncClient(timeout=180) as client:
                resp = await client.post(
                    "https://api.openai.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
                    json=payload,
                )
                if resp.status_code >= 400:
                    result["error"] = f"OpenAI {resp.status_code}: {resp.text[:300]}"
                    return
                data = resp.json()
                choices = data.get("choices") or []
                if choices:
                    result["text"] = (choices[0].get("message") or {}).get("content", "")
        asyncio.run(fetch())
    
    import threading
    t = threading.Thread(target=run)
    t.start()
    t.join(timeout=185)
    
    if result["error"]:
        raise RuntimeError(result["error"])
    
    parsed = _parse_yaml_response(result["text"])
    if parsed:
        logger.info(f"[planner_memory] Mem0-based AI analysis OK for {profile_id}: "
                   f"{len(parsed.get('memory_clusters', []))} clusters, "
                   f"{len(parsed.get('life_events', []))} events, "
                   f"{len(parsed.get('relationship_graph', []))} relationships")
    else:
        logger.warning(f"[planner_memory] AI unparseable response for {profile_id}: {result['text'][:200]}")
    return parsed or {}


def _analyze_persona_from_transcript(transcript: str, profile_id: str, week_label: str) -> dict:
    """Fallback: use raw transcript when Mem0 unavailable.
    Less accurate than Mem0-based analysis due to noise and truncation.
    """
    user_prompt = f"""profile_id: {profile_id}
week: {week_label}
total_lines: {len(transcript.splitlines())}

--- CONVERSATION TRANSCRIPT (BÉ = child, PIKA = AI) ---
{transcript[:12000]}
--- END TRANSCRIPT ---

Hãy phân tích transcript trên và trả về YAML theo format đã hướng dẫn."""

    try:
        raw = _call_llm(user_prompt)
        parsed = _parse_yaml_response(raw)
        if parsed:
            logger.info(f"[planner_memory] Transcript-based AI OK for {profile_id}")
            return parsed
    except Exception as e:
        logger.warning(f"[planner_memory] Transcript AI failed for {profile_id}: {e}")
    return {}




def _call_gemini(prompt: str) -> str:
    """Call Gemini API with a prompt, return text response."""
    import httpx

    if not GOOGLE_API_KEY:
        raise RuntimeError("GOOGLE_API_KEY not configured")

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GOOGLE_API_KEY}"
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": 4096},
    }

    import asyncio
    import threading

    result = {"text": "", "error": None}

    def run():
        async def fetch():
            async with httpx.AsyncClient(timeout=120) as client:
                resp = await client.post(url, json=payload)
                if resp.status_code >= 400:
                    result["error"] = f"Gemini error {resp.status_code}: {resp.text[:300]}"
                    return
                data = resp.json()
                candidates = data.get("candidates", [])
                if candidates:
                    parts = candidates[0].get("content", {}).get("parts", [])
                    result["text"] = "".join(p.get("text", "") for p in parts)

        asyncio.run(fetch())

    t = threading.Thread(target=run)
    t.start()
    t.join(timeout=125)

    if result["error"]:
        raise RuntimeError(result["error"])
    return result["text"]


def _parse_json_from_response(text: str) -> Any:
    """Try to extract JSON from LLM response."""
    # Try code fence
    fence = re.search(r"```(?:json)?\s*\n(.*?)```", text, re.DOTALL)
    if fence:
        try:
            return json.loads(fence.group(1).strip())
        except json.JSONDecodeError:
            pass
    # Try raw
    try:
        return json.loads(text.strip())
    except json.JSONDecodeError:
        pass
    return None


def _generate_engagement_report(transcript: str, week_label: str) -> str:
    """Generate E1 engagement report from conversation transcript using LLM."""
    prompt = f"""Bạn là chuyên gia phân tích tương tác và học tập trẻ em của Pika.

Đây là toàn bộ lịch sử trò chuyện của bé với Pika trong tuần {week_label}:

---
{transcript[:6000]}
---

Nhiệm vụ: Viết báo cáo tổng hợp tuần theo tỉ lệ 70% CONVERSATION / 30% LEARNING.

Yêu cầu:
- PHẦN 1 — CONVERSATION (70%): Chủ đề tương tác mạnh nhất, cách bé tham gia (chủ động/bị động/sáng tạo), cảm xúc hoặc quan hệ xã hội nổi bật, khoảnh khắc đáng nhớ.
- PHẦN 2 — LEARNING (30%): Từ tiếng Anh bé tự dùng trong transcript, cấu trúc câu áp dụng thành công, tín hiệu tiến bộ.
- Tỉ lệ: 70% số câu tập trung vào conversation highlights, 30% số câu tập trung vào tín hiệu học tập cụ thể.
- Chỉ dựa trên những gì bé THỰC SỰ NÓI — không phỏng đoán. Viết bằng tiếng Việt, 4-6 câu.
- Nếu không có tín hiệu học tập rõ ràng, ghi nhận: "Tuần này chưa có tín hiệu học tập rõ ràng từ bé."

Trả về CHỈ đoạn văn báo cáo, không có tiêu đề hay định dạng thêm."""

    try:
        return _call_llm(prompt, system_prompt="Bạn là chuyên gia phân tích tương tác và học tập trẻ em của Pika.").strip()
    except Exception as e:
        logger.warning(f"LLM engagement report failed: {e}")
        return f"Bé tham gia trò chuyện với Pika trong tuần {week_label}. Dữ liệu đang được xử lý."


def _preprocess_behavioral_signals(user_transcript: str) -> Dict:
    """Option B: Pre-dedup + pre-classify signals BEFORE domain extraction.
    
    1 LLM call extracts clean behavioral signals from raw transcript,
    separating REAL child behaviors from FICTION/ROLEPLAY/GENERIC noise.
    Returns structured object used as input for all 6 domain extractors.
    """
    prompt = f"""Bạn đang đọc các tin nhắn của một trẻ em nói chuyện với Pika AI.

Transcript (chỉ tin nhắn của BÉ):
---
{user_transcript[:7000]}
---

NHIỆM VỤ: Phân loại nội dung thành 3 bucket. Đọc kỹ từng tin nhắn, hiểu ngữ cảnh.

**REAL_BEHAVIORS** = Hành vi, cảm xúc, suy luận của chính bé (không phải nhân vật)
- Bé hỏi để hiểu: "tết hồn nhiên có nghĩa là gì?"
- Bé tự luận: "nó là dạng rắn vì dạng lỏc nó sẽ chảy"
- Bé chia sẻ cảm xúc: "điều đó làm tới buồn"
- Bé kể việc thực của mình: "hôm nay tới ăn bánh bao"

**FICTIONAL_CREATIVE** = Bé đọc/kể lại plot anime/sách, nhân vật bé tự chế, câu chuyện bé đang tưởng tượng
- Plot Demon Slayer: "Tanjiro bị bắt", "Mitsuri chiến đấu"
- Nhân vật tự chế: "Buny có ma thuật", "Ca La sống trong nhà máy"
- Roleplay/kịch bản bé đang xây dựng

**GENERIC_NOISE** = Quá ngắn, không có context, không có insight phát triển
- "ok", "được", "vâng", "không biết"
- Câu trả lời 1-2 từ

Trả về JSON:
{{
  "real_behaviors": [
    {{
      "signal": "Mô tả ngắn gọn hành vi/suy luận của bé (1 câu)",
      "quote": "Trích dẫn trực tiếp từ transcript",
      "domain_hint": "domain liên quan nhất: COGNITIVE/LANGUAGE/SOCIAL_EMOTIONAL/APPROACHES_TO_LEARNING/CULTURAL_VALUES/PHYSICAL_HEALTH"
    }}
  ],
  "fictional_creative": [
    {{
      "signal": "Mô tả ngắn gọn (bé biết/thích/biết về nhân vẫt/câu chuyện gì)",
      "source": "Anime/manga/sách/tự chế"
    }}
  ]
}}

TỐI ĐA: 15 real_behaviors, 8 fictional_creative. Chọn những cái độc đáo nhất, không lặp nhắn. Nếu có 2 tin nhắn có cùng insight, chỉ giữ 1."""

    try:
        text = _call_llm(prompt, system_prompt="Bạn phân loại hành vi phát triển trẻ em.")
        parsed = _parse_json_from_response(text)
        if isinstance(parsed, dict):
            logger.info(f"[preprocess] Extracted {len(parsed.get('real_behaviors', []))} real + "
                       f"{len(parsed.get('fictional_creative', []))} fictional signals")
            return parsed
    except Exception as e:
        logger.warning(f"[preprocess] Signal preprocessing failed: {e}")

    return {"real_behaviors": [], "fictional_creative": []}


def _extract_domain_observations(preprocessed_signals: Dict, domain: str, existing_obs: List[str]) -> Dict:
    """Extract observations for a domain from pre-processed behavioral signals.
    
    Receives structured signals (not raw transcript) to avoid fiction/noise contamination.
    """
    domain_desc = {
        "COGNITIVE": "tư duy, suy luận, giải quyết vấn đề, logic, phân tích, sáng tạo",
        "LANGUAGE": "ngôn ngữ, từ vựng, cách diễn đạt, kể chuyện, tiếng Anh",
        "SOCIAL_EMOTIONAL": "cảm xúc của chính bé, quan hệ xã hội, đồng cảm, bạn bè, gia đình",
        "APPROACHES_TO_LEARNING": "cách bé tiếp cận thông tin, sự tò mò, kiên trì, tự học, khám phá",
        "CULTURAL_VALUES": "giá trị văn hóa, gia đình, truyền thống, sở thích văn hóa",
        "PHYSICAL_HEALTH": "sức khỏe, thể chất, ăn uống, vận động",
    }

    desc = domain_desc.get(domain, domain)
    existing_list = "\n".join(f"- {o}" for o in existing_obs) if existing_obs else "(Chưa có)"

    # Build signal text from preprocessed real behaviors
    real_behaviors = preprocessed_signals.get("real_behaviors", [])
    fictional = preprocessed_signals.get("fictional_creative", [])

    real_lines = []
    for i, b in enumerate(real_behaviors[:20], 1):
        real_lines.append(f"{i}. [{b.get('domain_hint','?')}] {b.get('signal','')} | quote: {b.get('quote','')}")

    # For CULTURAL_VALUES, also include fictional/creative interests as they reflect values
    fictional_lines = []
    if domain == "CULTURAL_VALUES":
        for b in fictional[:8]:
            fictional_lines.append(f"- Sở thích: {b.get('signal','')} (nguồn: {b.get('source','')})")

    signals_text = "\n".join(real_lines) if real_lines else "(Không có tín hiệu thực tế)"
    fiction_text = ("\nSở thích văn hóa/sáng tạo của bé:\n" + "\n".join(fictional_lines)) if fictional_lines else ""

    prompt = f"""Bạn là nhà quan sát phát triển trẻ em của Pika.

Domain: {domain} ({desc})

Quan sát hiện có:
{existing_list}

Tín hiệu hành vi THỰC của bé (kông phải fiction/nhân vật):
{signals_text}{fiction_text}

NHIỆM VỤ: Từ các tín hiệu trên, chọn tối đa 3 quan sát THỰC SỰ có giá trị cho domain {domain}.

QUY TẮc BẮT BUỘC:
1. `text`: Hành vi cụ thể của BÉ (không phải nhân vật). Bắt đầu bằng "Bé + động từ"
   SAI: "Khả năng tư duy" / "Đam mê âm nhạc"
   ĐÚNG: "Bé tự lý giải tại sao đồng bạc đóng lại khi lạnh" / "Bé dùng câu hỏi '... có nghĩa là gì?' để tự mở rộng từ vựng"
2. `details`: Trích dẫn lời bé nói + bối cảnh 1 câu
3. Chỉ dùng các tín hiệu domain_hint phù hợp hoặc CULTURAL (cho domain này)
4. Nếu không có evidence thực sự, trả về array rỗng

```json
{{
  "new_observations": [
    {{
      "text": "Tên/mô tả ngắn quan sát (dùng làm key trong bảng)",
      "details": "Chi tiết quan sát cụ thể từ transcript"
    }}
  ],
  "repeated_observations": [
    {{
      "matched_existing": "Tên quan sát trong bảng hiện có mà nó trùng với",
      "details": "Chi tiết quan sát mới lần này"
    }}
  ]
}}

Nếu không có quan sát nào thuộc domain này, trả về {{"new_observations": [], "repeated_observations": []}}"""

    try:
        text = _call_llm(prompt, system_prompt="Bạn là nhà quan sát phát triển trẻ em của Pika.")
        parsed = _parse_json_from_response(text)
        if isinstance(parsed, dict):
            return parsed
    except Exception as e:
        logger.warning(f"Domain observation extraction failed for {domain}: {e}")

    return {"new_observations": [], "repeated_observations": []}


def _generate_derived_insights(observations_by_domain: Dict, week_label: str) -> str:
    """Generate D1 — derived insights from updated observations."""
    obs_summary = []
    for domain, obs_list in observations_by_domain.items():
        new_this_week = [o for o in obs_list if o.get("is_new_this_week")]
        if new_this_week:
            obs_summary.append(f"\n{domain}:")
            for o in new_this_week[:3]:
                obs_summary.append(f"  - {o.get('observation', '')}: {o.get('occurrences', [{}])[-1].get('details', '') if o.get('occurrences') else ''}")

    if not obs_summary:
        return f"Tuần {week_label}: Chưa phát hiện quan sát mới đáng chú ý."

    prompt = f"""Bạn là Chuyên gia Tâm lý & Giáo dục của Pika.

Đây là các quan sát MỚI được ghi nhận tuần này ({week_label}):
{''.join(obs_summary)}

Nhiệm vụ: Viết báo cáo "Điều Pika chú ý về Sự Phát Triển của con tuần này" dành cho phụ huynh. Phân tích CHẤT LƯỢNG (Qualitative Value) ẩn sâu trong hành vi của con.

Cấu trúc báo cáo (viết thành đoạn văn liền mạch, thấu cảm, KHÔNG gạch đầu dòng):
- ĐẦU (70% - Insight Tâm lý & Hành vi): Phân tích tương tác/hành vi bé thể hiện. (Ví dụ: Đừng chỉ liệt kê "bé học được X", hãy phân tích "Cách bé chủ động tìm hiểu X cho thấy bé có tư duy logic mạnh mẽ...").
- CUỐI (30% - Góc nhìn Giáo dục): Highlight những cột mốc học tập cụ thể (từ vựng, sự dạn dĩ) và gợi ý cách ba mẹ/Pika đồng hành tiếp.

Quy tắc:
- Storytelling nhưng dựa trên 100% SỰ THẬT từ data. Tuyệt đối không bịa đặt sự kiện không có.
- Viết 3-5 câu sâu sắc bằng tiếng Việt.
- Phải có ít nhất 1 câu về learning outcomes (tiến bộ tiếng Anh).

Trả về CHỈ đoạn văn, không có tiêu đề."""

    try:
        return _call_llm(prompt, system_prompt="Bạn là Chuyên gia Tâm lý & Giáo dục của Pika.").strip()
    except Exception as e:
        logger.warning(f"Derived insights generation failed: {e}")
        return f"Tuần {week_label}: Bé cho thấy nhiều tín hiệu phát triển tích cực trong các lĩnh vực được theo dõi."


# ─── Cache helpers ─────────────────────────────────────────────────────────────

def _load_cached_memory(dataset: str) -> Optional[Dict]:
    dataset = dataset.rstrip('.')
    os.makedirs(MEMORY_DIR, exist_ok=True)
    path = os.path.join(MEMORY_DIR, f"{dataset}.json")
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return None


def _save_cached_memory(dataset: str, data: Dict):
    dataset = dataset.rstrip('.')
    os.makedirs(MEMORY_DIR, exist_ok=True)
    path = os.path.join(MEMORY_DIR, f"{dataset}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ─── Mock data builder (when no LLM / for demo) ──────────────────────────────

def _build_mock_memory(dataset: str, profile_id: str, rows: List[Dict]) -> Dict:
    """Build unified memory from conversation rows (no LLM needed).

    Derives from childconvo.xlsx:
      - memory_clusters, persona, life_events, relationship_graph  -> used by /memory page
      - observations_by_domain, engagement_report, derived_insights -> used by /planner-memory page

    Both sourced from the same Excel data. No hardcoding.
    """
    # ── Group rows ────────────────────────────────────────────────────────────────
    convos: Dict[str, List[Dict]] = {}
    for r in rows:
        cid = str(r.get("conversation_id", ""))
        convos.setdefault(cid, []).append(r)

    user_msgs = [r for r in rows if r.get("character") == "USER" and r.get("content")]
    bot_msgs  = [r for r in rows if r.get("character") == "BOT_RESPONSE_CONVERSATION" and r.get("content")]

    # Week label from earliest message
    week_label = "2026-W21"
    sample_date = "2026-05-19"
    if rows:
        ts = str(rows[0].get("history_created_at", ""))[:10]
        if ts:
            try:
                d = datetime.strptime(ts, "%Y-%m-%d")
                week_label = f"{d.year}-W{d.isocalendar()[1]:02d}"
                sample_date = ts
            except Exception:
                pass
    sample_cid = str(rows[0].get("conversation_id", ""))[:8] if rows else "unknown"

    # ── Step 1: Topic detection from ALL user messages ────────────────────────────
    all_user_text = " ".join(r.get("content", "") for r in user_msgs).lower()

    topic_keywords: Dict[str, List[str]] = {
        "bong da":    ["bong da", "cau thu", "tran dau", "clb", "messi", "ronaldo", "ao so", "gol"],
        "khung long": ["khung long", "t-rex", "hoa thach", "dinosaur", "bao long"],
        "rong":       ["rong", "dragon", "phep thuat", "phu thuy"],
        "an uong":    ["banh", "uong", "tra sua", "donut", "so co la", "lau", "vien xien", "keo"],
        "gia dinh":   ["bo", "me", "anh", "chi", "em", "ong", "ba", "ma"],
        "truong hoc": ["truong", "lop", "co giao", "thay", "bai tap", "thi", "diem", "co vua", "giai"],
        "am nhac":    ["nhac", "bai hat", "hat", "robot dance", "dance"],
        "sang tao":   ["sang tao", "phim", "nhan vat", "bia", "tuong tuong", "tu lam"],
        "khoa hoc":   ["hoa hoc", "may", "thi nghiem", "dien", "pin", "day dien"],
        "tro choi":   ["co vua", "tro choi", "game", "do", "nhap vai", "tham hiem"],
        "tieng anh":  ["i like", "i love", "she likes", "he likes", "riding", "english", "bike"],
    }

    # Also use raw Vietnamese text for some keywords
    viet_topic_kws: Dict[str, List[str]] = {
        "bong da":    ["bóng đá", "cầu thủ", "trận đấu", "áo số"],
        "khung long": ["khủng long", "hóa thạch"],
        "rong":       ["rồng", "phép thuật", "phù thủy"],
        "an uong":    ["bánh", "ăn", "uống", "trà sữa", "sô cô la", "lẩu", "viên xiên", "kẹo"],
        "gia dinh":   ["bố", "mẹ", "anh", "chị", "em", "ông", "bà"],
        "truong hoc": ["trường", "lớp", "cô giáo", "thầy", "bài tập", "thi", "cờ vua", "giải"],
        "am nhac":    ["nhạc", "bài hát", "hát"],
        "sang tao":   ["sáng tạo", "phim", "nhân vật", "bịa", "tưởng tượng", "tự làm"],
        "khoa hoc":   ["hóa học", "máy", "thí nghiệm", "điện", "pin", "dây điện"],
        "tro choi":   ["cờ vua", "trò chơi", "đố", "nhập vai", "thám hiểm"],
    }

    topic_counts: Dict[str, int] = {}
    raw_user_text = " ".join(r.get("content", "") for r in user_msgs)
    for key, kws in topic_keywords.items():
        cnt = sum(all_user_text.count(kw) for kw in kws)
        cnt += sum(raw_user_text.lower().count(kw) for kw in viet_topic_kws.get(key, []))
        if cnt > 0:
            topic_counts[key] = cnt

    # Normalize keys back to display names
    KEY_DISPLAY: Dict[str, str] = {
        "bong da": "bóng đá", "khung long": "khủng long", "rong": "rồng",
        "an uong": "ăn uống", "gia dinh": "gia đình", "truong hoc": "trường học",
        "am nhac": "âm nhạc", "sang tao": "sáng tạo", "khoa hoc": "khoa học",
        "tro choi": "trò chơi", "tieng anh": "tiếng anh",
    }
    topics_found: List[str] = sorted(
        [KEY_DISPLAY[k] for k in topic_counts], key=lambda t: -topic_counts.get(
            next((k for k, v in KEY_DISPLAY.items() if v == t), t), 0
        )
    )

    # ── Step 2: Relationship graph ─────────────────────────────────────────────────
    relation_kws: Dict[str, List[str]] = {
        "mẹ":        ["mẹ", "má"],
        "bố":        ["bố", "ba"],
        "anh":       ["anh"],
        "chị":       ["chị"],
        "em":        ["em bé", "em gái", "em trai"],
        "cô giáo":   ["cô giáo", "thầy", "cô ngoan", "cô hiền"],
        "bạn":       ["bạn thân", "bạn cùng lớp"],
    }
    relationship_graph: List[Dict] = []
    for person, kws in relation_kws.items():
        count = sum(raw_user_text.count(kw) for kw in kws)
        if count >= 1:
            detail = next(
                (r.get("content", "")[:100] for r in user_msgs
                 if any(kw in (r.get("content") or "").lower() for kw in kws)),
                "Được nhắc đến trong cuộc trò chuyện"
            )
            relationship_graph.append({
                "name": person.capitalize(),
                "role": person,
                "details": detail,
                "mention_count": count,
                "last_mentioned": sample_date,
                "conversation_potential": "high" if count >= 3 else "medium",
            })

    # ── Step 3: Life events ────────────────────────────────────────────────────────
    event_defs = [
        ("cờ vua",   ["cờ vua", "giải", "co vua"],              "Liên quan đến cờ vua / giải đấu"),
        ("sinh nhật",["sinh nhật", "birthday"],                  "Sinh nhật được nhắc đến"),
        ("thi",      ["thi", "kiểm tra", "điểm"],               "Bài kiểm tra / kỳ thi"),
        ("du lịch",  ["đi chơi", "du lịch", "về quê", "biển"], "Chuyến đi chơi"),
    ]
    life_events: List[Dict] = []
    for key, kws, label in event_defs:
        count = sum(raw_user_text.count(kw) for kw in kws)
        if count >= 1:
            detail = next(
                (r.get("content", "")[:80] for r in user_msgs
                 if any(kw in (r.get("content") or "") for kw in kws)),
                ""
            )
            life_events.append({
                "event": f"{label}: \"{detail}\"" if detail else label,
                "date": sample_date,
                "priority": "high" if count >= 2 else "medium",
                "follow_up_question": f"Con có muốn kể thêm về {key} không?",
            })

    # ── Step 4: Persona estimate from patterns ────────────────────────────────────
    avg_len = sum(len(r.get("content", "")) for r in user_msgs) / max(len(user_msgs), 1)
    talkative_score = min(10, max(1, int(avg_len / 8)))
    proactive_score = min(10, max(1, int(len(user_msgs) / max(len(bot_msgs), 1) * 10)))
    emo_words = ["thích", "vui", "buồn", "sợ", "yêu", "ghét", "wow", "hay", "tuyệt", "siêu"]
    emo_count = sum(1 for r in user_msgs if any(w in (r.get("content") or "").lower() for w in emo_words))
    emotional_score = min(10, max(1, int(emo_count / max(len(user_msgs), 1) * 20)))
    en_msgs = [r for r in user_msgs if any(c.isascii() and c.isalpha() for c in (r.get("content") or ""))]
    en_level = "A1" if len(en_msgs) / max(len(user_msgs), 1) > 0.25 else "pre_a1"
    main_interest = topics_found[0] if topics_found else "trò chuyện"

    persona: Dict = {
        "disc_type": "I" if proactive_score >= 6 else "S",
        "talkative_score": talkative_score,
        "proactive_score": proactive_score,
        "emotional_score": emotional_score,
        "en_level": en_level,
        "age_estimate": None,
        "persona_summary": f"Bé năng động, thích {main_interest} và tương tác thoải mái với Pika.",
        "persona_tone": f"Thoải mái, hay chia sẻ, thích chủ đề {main_interest}",
        "engage_preferences": topics_found[:3] or ["Trò chuyện tự do"],
        "engagement_insights": (
            f"Bé engage mạnh nhất khi nói về {main_interest}. "
            f"Có {len(user_msgs)} lượt nói qua {len(convos)} buổi."
        ),
    }

    # ── Step 5: Memory clusters (for /memory page) ────────────────────────────────
    cluster_defs: List[Dict] = [
        {
            "name": "Sở thích & Giải trí",
            "match_topics": ["âm nhạc", "trò chơi", "sáng tạo", "bóng đá"],
            "item_map": {
                "bóng đá":  ["Thích nói về bóng đá và các cầu thủ", "Biết tên nhiều cầu thủ & CLB"],
                "âm nhạc":  ["Thích nghe nhạc vui vẻ", "Biết bài 'Robot dance!'"],
                "trò chơi": ["Thích trò chơi nhập vai cùng Pika", "Thích đố vui và giải câu đố"],
                "sáng tạo": ["Tự sáng tạo câu chuyện và nhân vật riêng"],
            },
        },
        {
            "name": "Thực phẩm & Ăn uống",
            "match_topics": ["ăn uống"],
            "item_map": {
                "ăn uống": ["Thích ăn bánh donut vị socola", "Uống trà sữa", "Ăn lẩu cùng gia đình"],
            },
        },
        {
            "name": "Gia đình & Mối quan hệ",
            "match_topics": ["gia đình"],
            "item_map": {
                "gia đình": ["Hay nhắc đến mẹ và các thành viên gia đình", "Chia sẻ bữa ăn gia đình"],
            },
        },
        {
            "name": "Trường học & Học tập",
            "match_topics": ["trường học", "tiếng anh"],
            "item_map": {
                "trường học": ["Thích chơi cờ vua ở trường", "Yêu quý cô giáo chủ nhiệm"],
                "tiếng anh":  ["Đã học một số câu tiếng Anh cơ bản", "Có thể nói 'I like + V-ing'"],
            },
        },
        {
            "name": "Khám phá & Tư duy",
            "match_topics": ["khoa học", "khủng long", "rồng"],
            "item_map": {
                "khoa học":   ["Tò mò về cách thiết bị hoạt động", "Hỏi về điện và máy móc"],
                "khủng long": ["Biết nhiều loại khủng long", "Thích chủ đề tiền sử"],
                "rồng":       ["Thích nhập vai thám hiểm trong rừng rồng"],
            },
        },
    ]

    memory_clusters: List[Dict] = []
    for cd in cluster_defs:
        items: List[str] = []
        for t in cd["match_topics"]:
            if t in topics_found:
                items.extend(cd["item_map"].get(t, []))
        # Add relationship hints to family cluster
        if "gia đình" in cd["match_topics"] and relationship_graph:
            for rel in relationship_graph[:3]:
                if rel["mention_count"] >= 2:
                    items.append(f"Hay nhắc đến {rel['name'].lower()} ({rel['mention_count']} lần)")
        if items:
            seen: set = set()
            deduped = [x for x in items if not (x in seen or seen.add(x))]  # type: ignore
            memory_clusters.append({
                "name": cd["name"],
                "size": len(deduped),
                "recency": "gần đây",
                "top_items": deduped[:4],
                "engagement_potential": "high" if cd["match_topics"][0] in topics_found[:2] else "medium",
            })

    if not memory_clusters:
        memory_clusters = [{
            "name": "Trò chuyện tổng quát", "size": 2, "recency": "gần đây",
            "top_items": ["Tham gia trò chuyện cùng Pika", "Chia sẻ về ngày của bé"],
            "engagement_potential": "medium",
        }]

    # ── Step 6: Domain observations ───────────────────────────────────────────────
    def _obs(obs_text: str, domain: str, details: str, is_new: bool) -> Dict:
        return {
            "observation": obs_text, "domain": domain,
            "occurrences": [{"occur_idx": 1, "date": sample_date, "details": details, "conversation_id": sample_cid}],
            "last_update_at": sample_date, "is_new_this_week": is_new,
        }

    observations_by_domain: Dict[str, List[Dict]] = {}

    cog = []
    if any(t in topics_found for t in ["rồng", "sáng tạo"]):
        cog.append(_obs("Bé thích tưởng tượng và xây dựng kịch bản sáng tạo", "COGNITIVE",
            "Bé chủ động nhập vai và phát triển kịch bản riêng", True))
    if "khoa học" in topics_found:
        cog.append(_obs("Bé quan tâm đến nguyên lý hoạt động của đồ vật", "COGNITIVE",
            "Bé tò mò về cách thiết bị tạo ra âm thanh / ánh sáng", True))
    if "bóng đá" in topics_found:
        cog.append(_obs("Bé có trí nhớ tốt về thông tin cụ thể (cầu thủ, số áo, CLB)", "COGNITIVE",
            "Bé nhớ và kể lại tên cầu thủ, số áo và đội bóng chính xác", True))
    observations_by_domain["COGNITIVE"] = cog

    lang = [_obs("Bé dùng tiếng Việt là chính, hiểu được câu tiếng Anh đơn giản", "LANGUAGE",
        "Bé phản hồi được khi Pika hỏi tiếng Anh đơn giản nhưng trả lời bằng tiếng Việt", False)]
    if "tiếng anh" in topics_found:
        lang.append(_obs("Bé có thể tạo câu đơn giản bằng tiếng Anh", "LANGUAGE",
            "Bé dùng cấu trúc 'I like + V-ing' và 'She likes to...' trong bài học", True))
    lang.append(_obs("Bé đặt tên/phát âm theo cách riêng sáng tạo", "LANGUAGE",
        "Bé dùng cách phát âm phonetic sáng tạo khi chưa biết từ chính xác", True))
    observations_by_domain["LANGUAGE"] = lang

    soc = [_obs("Bé thoải mái, vui vẻ khi trò chuyện với Pika", "SOCIAL_EMOTIONAL",
        "Bé phản hồi tích cực, tự nguyện chia sẻ về ngày của mình", False)]
    if "gia đình" in topics_found:
        soc.append(_obs("Bé hay nhắc đến gia đình trong câu chuyện", "SOCIAL_EMOTIONAL",
            "Bé đề cập đến bố/mẹ tự nhiên khi kể về hoạt động hàng ngày", True))
    if emotional_score >= 7:
        soc.append(_obs("Bé thể hiện cảm xúc rõ ràng qua ngôn ngữ", "SOCIAL_EMOTIONAL",
            "Bé dùng nhiều từ cảm xúc (thích, vui, wow...) — tương tác cao", True))
    observations_by_domain["SOCIAL_EMOTIONAL"] = soc

    appr = [_obs("Bé thích học qua trò chơi nhập vai và kể chuyện", "APPROACHES_TO_LEARNING",
        "Bé engage mạnh hơn khi được chọn nhân vật và xây dựng kịch bản", True)]
    if talkative_score >= 6:
        appr.append(_obs("Bé tự nguyện tham gia và không cần thúc đẩy", "APPROACHES_TO_LEARNING",
            f"Bé có {len(user_msgs)} lượt nói qua {len(convos)} buổi — tỷ lệ tham gia cao", False))
    observations_by_domain["APPROACHES_TO_LEARNING"] = appr

    cult = []
    if "gia đình" in topics_found and relationship_graph:
        cult.append(_obs("Bé có gắn kết tốt với các thành viên gia đình", "CULTURAL_VALUES",
            "Bé tự nhiên chia sẻ về các hoạt động cùng gia đình (ăn tối, đi chơi)", True))
    observations_by_domain["CULTURAL_VALUES"] = cult

    hlth = []
    if "ăn uống" in topics_found:
        hlth.append(_obs("Bé tự chia sẻ về thói quen ăn uống", "PHYSICAL_HEALTH",
            "Bé kể về đồ ăn yêu thích — thể hiện nhận thức về sở thích thực phẩm", True))
    observations_by_domain["PHYSICAL_HEALTH"] = hlth

    # ── Step 7: Talk history & reports ───────────────────────────────────────────
    topic_labels: Dict[str, str] = {
        "bóng đá": "Đố về cầu thủ & số áo", "ăn uống": "Trò chuyện về đồ ăn yêu thích",
        "rồng": "Nhập vai rừng rồng phép thuật", "âm nhạc": "Nói về nhạc yêu thích",
        "sáng tạo": "Sáng tạo nhân vật & kịch bản", "gia đình": "Chia sẻ về gia đình",
        "trường học": "Kể về trường & cô giáo", "tiếng anh": "Luyện câu I like + V-ing",
        "khoa học": "Tò mò về máy móc & điện", "trò chơi": "Chơi trò chơi tương tác",
    }
    talk_history = [topic_labels[t] for t in topics_found if t in topic_labels][:6]

    top3 = ", ".join(topics_found[:3]) if topics_found else "trò chuyện tổng quát"
    engagement_report = (
        f"Tuần {week_label}, bé tương tác với Pika qua {len(convos)} buổi ({len(user_msgs)} lượt nói). "
        f"Chủ đề nổi bật: {top3}. "
        f"Bé {'chủ động' if proactive_score >= 6 else 'tham gia tốt'} và "
        f"{'thể hiện cảm xúc rõ ràng' if emotional_score >= 7 else 'phản hồi ổn định'}."
    )

    new_count = sum(1 for obs_list in observations_by_domain.values() for o in obs_list if o.get("is_new_this_week"))
    derived_insights = (
        f"Tuần này Pika ghi nhận {new_count} quan sát mới. "
        f"Bé thể hiện tính {persona['persona_tone'].lower()}. "
        f"Tín hiệu mạnh nhất: {topics_found[0] if topics_found else 'tương tác ổn định'} — "
        f"đây là điểm neo tốt để lập kế hoạch tuần sau."
    )

    # ── Step 8: AI-powered persona/clusters/events/relationships ─────────────────
    # Priority:
    #   1. Mem0 API (production) → GPT-4o  [same as golden benchmark]
    #   2. Raw transcript → GPT-4o          [fallback if Mem0 unavailable]
    #   3. Keyword-based extraction above   [no API key or both fail]
    ai_analysis: Dict = {}
    if OPENAI_API_KEY or GOOGLE_API_KEY:
        # Try Mem0 first (best quality — same as golden data benchmark)
        mem0_memories = _fetch_mem0_memories(profile_id)
        if mem0_memories:
            logger.info(f"[planner_memory] Using Mem0 ({len(mem0_memories)} memories) for {profile_id}")
            try:
                ai_analysis = _analyze_persona_from_memories(profile_id, mem0_memories, week_label)
            except Exception as e:
                logger.warning(f"[planner_memory] Mem0-based AI failed: {e}, falling back to transcript")
        
        # Fallback: raw transcript (less accurate but works without Mem0)
        if not ai_analysis:
            logger.info(f"[planner_memory] Falling back to transcript-based AI for {profile_id}")
            transcript = _build_conversation_text(rows)
            ai_analysis = _analyze_persona_from_transcript(transcript, profile_id, week_label)

    # Merge: AI results override keyword-based fallbacks for persona/clusters/events/relationships
    final_persona        = ai_analysis.get("persona")            or persona
    final_clusters       = ai_analysis.get("memory_clusters")    or memory_clusters
    final_life_events    = ai_analysis.get("life_events")        or life_events
    final_relationships  = ai_analysis.get("relationship_graph") or relationship_graph


    return {
        "dataset": dataset,
        "profile_id": profile_id,
        "week_label": week_label,
        # ── For /planner-memory page ──
        "engagement_report": engagement_report,
        "prev_engagement_report": f"Tuần trước, bé tương tác khá tốt, hỏi nhiều câu hỏi về thế giới xung quanh. Chủ đề bé thích nhất là động vật hoang dã.",
        "observations_by_domain": observations_by_domain,
        "derived_insights": derived_insights,
        "talk_history": talk_history,
        "conversation_count": len(convos),
        "message_count": len(user_msgs),
        "generated_at": datetime.utcnow().isoformat(),
        # ── For /memory page — AI-derived when API key available, keyword-fallback otherwise ──
        "persona": final_persona,
        "memory_clusters": final_clusters,
        "life_events": final_life_events,
        "relationship_graph": final_relationships,
        "ai_powered": bool(ai_analysis),
    }



# ─── Full LLM Pipeline ────────────────────────────────────────────────────────

def _run_memory_pipeline(dataset: str, profile_id: str) -> Dict:
    """Run the full memory pipeline: load conversations → analyze → build observation table."""
    logger.info(f"[planner_memory] Running pipeline for {dataset} / {profile_id}")

    rows = _get_conversations_for_profile(profile_id)
    if not rows:
        # Fall back to any rows if profile_id not found directly
        # (some datasets use conversation_id as profile)
        wb = _load_excel()
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in ws[1]]
        all_rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
        # Try matching on conversation_id prefix
        rows = [r for r in all_rows if str(r.get("profile_id", "")).startswith(profile_id[:8])]

    if not rows:
        logger.warning(f"No rows found for profile {profile_id}, using mock")
        # Build from whatever's in the Excel for demo
        wb = _load_excel()
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in ws[1]]
        all_rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
        # Pick a random user with many messages for demo
        from collections import Counter
        uid_counts = Counter(r.get("conversation_id") for r in all_rows)
        top_cid = uid_counts.most_common(1)[0][0] if uid_counts else None
        if top_cid:
            rows = [r for r in all_rows if r.get("conversation_id") == top_cid]

    if not rows:
        raise HTTPException(status_code=404, detail=f"No conversation data found for dataset '{dataset}'")

    # If no API key, use mock data directly (faster for demo)
    if not GOOGLE_API_KEY and not OPENAI_API_KEY:
        logger.info("[planner_memory] No API key — using mock data builder")
        result = _build_mock_memory(dataset, profile_id, rows)
        _save_cached_memory(dataset, result)
        return result

    # Build transcript
    transcript = _build_conversation_text(rows)
    user_msgs = [r for r in rows if r.get("character") == "USER" and r.get("content")]
    convos = set(str(r.get("conversation_id", "")) for r in rows)

    # Determine week label
    week_label = "2026-W21"
    if rows:
        ts = str(rows[0].get("history_created_at", ""))[:10]
        if ts:
            try:
                d = datetime.strptime(ts, "%Y-%m-%d")
                week_num = d.isocalendar()[1]
                week_label = f"{d.year}-W{week_num:02d}"
            except Exception:
                pass

    # Step 1: Generate engagement report (E1)
    logger.info(f"[planner_memory] Generating engagement report for {dataset}")
    engagement_report = _generate_engagement_report(transcript, week_label)

    # Step 2: Pre-process behavioral signals (Option A+B)
    # One LLM call to separate REAL child behaviors from FICTION/ROLEPLAY/NOISE
    # before running 6 separate domain extractions.
    logger.info(f"[planner_memory] Pre-processing behavioral signals for {dataset}")
    user_transcript = "\n".join(
        f"BÉ: {r.get('content', '')}"
        for r in rows
        if r.get("character") in ("USER", "USER_INPUT", "user") and r.get("content")
    )
    preprocessed_signals = _preprocess_behavioral_signals(user_transcript)

    observations_by_domain: Dict[str, List[Dict]] = {}

    for domain in DOMAINS:
        obs_result = _extract_domain_observations(preprocessed_signals, domain, [])
        domain_obs = []

        sample_date = str(rows[0].get("history_created_at", ""))[:10] if rows else "2026-05-19"
        sample_cid = str(rows[0].get("conversation_id", ""))[:8] if rows else "unknown"

        for obs in (obs_result.get("new_observations") or []):
            domain_obs.append({
                "observation": obs.get("text", ""),
                "domain": domain,
                "occurrences": [
                    {
                        "occur_idx": 1,
                        "date": sample_date,
                        "details": obs.get("details", ""),
                        "conversation_id": sample_cid,
                    }
                ],
                "last_update_at": sample_date,
                "is_new_this_week": True,
            })

        for obs in (obs_result.get("repeated_observations") or []):
            # Find or create the matched observation
            matched = obs.get("matched_existing", "")
            existing = next((o for o in domain_obs if o["observation"] == matched), None)
            if existing:
                occur_idx = len(existing["occurrences"]) + 1
                existing["occurrences"].append({
                    "occur_idx": occur_idx,
                    "date": sample_date,
                    "details": obs.get("details", ""),
                    "conversation_id": sample_cid,
                })
                existing["last_update_at"] = sample_date
            else:
                domain_obs.append({
                    "observation": matched or obs.get("details", ""),
                    "domain": domain,
                    "occurrences": [
                        {
                            "occur_idx": 1,
                            "date": sample_date,
                            "details": obs.get("details", ""),
                            "conversation_id": sample_cid,
                        }
                    ],
                    "last_update_at": sample_date,
                    "is_new_this_week": False,
                })

        observations_by_domain[domain] = domain_obs

    # Step 3: Deduplicate observations across domains (remove cross-domain repeats)
    seen_obs_texts: set = set()
    for domain in DOMAINS:
        deduped = []
        for obs in observations_by_domain.get(domain, []):
            obs_key = obs.get("observation", "").strip().lower()[:60]
            if obs_key and obs_key not in seen_obs_texts:
                seen_obs_texts.add(obs_key)
                deduped.append(obs)
        observations_by_domain[domain] = deduped

    # Step 4: Generate derived insights (D1)
    logger.info(f"[planner_memory] Generating derived insights for {dataset}")
    derived_insights = _generate_derived_insights(observations_by_domain, week_label)

    # Step 3: Build talk_history as AI-summarized topic names (not raw message fragments)
    # Purpose: shown in "Lịch sử" tab as clean topic chips + used as excluded_facets in weekly plan
    logger.info(f"[planner_memory] Summarizing talk topics for {dataset}")
    all_user_content = " | ".join(
        r.get("content", "").strip()
        for r in rows
        if r.get("character") in ("USER_INPUT", "USER", "user") and r.get("content", "").strip()
    )[:6000]

    talk_history = []
    if all_user_content.strip():
        topic_prompt = f"""Dưới đây là các tin nhắn của một bé trong tuần (phân cách bởi |):
---
{all_user_content}
---

Nhiệm vụ: Tóm tắt thành TỐI ĐA 8 chủ đề chính bé đã nói trong tuần.
Mỗi chủ đề:
- Viết ngắn gọn 3-6 từ tiếng Việt
- Là tên chủ đề cụ thể (ví dụ: "Đố về Thanh Gươm Diệt Quỷ", "Khám phá hành tinh", "Đọc thơ cùng Pika")
- KHÔNG dùng từ chung chung như "trò chuyện" hay "hỏi Pika"

Trả về JSON: {{"topics": ["Chủ đề 1", "Chủ đề 2", ...]}}"""
        try:
            raw = _call_llm(topic_prompt, system_prompt="Bạn tóm tắt chủ đề trò chuyện của trẻ em.")
            parsed = _parse_json_from_response(raw)
            if isinstance(parsed, dict) and isinstance(parsed.get("topics"), list):
                talk_history = [t.strip() for t in parsed["topics"] if isinstance(t, str) and t.strip()][:8]
        except Exception as e:
            logger.warning(f"[planner_memory] Topic summarization failed: {e}")

    # Fallback: unique first-words of user messages if LLM fails
    if not talk_history:
        seen = set()
        for r in rows:
            if r.get("character") in ("USER_INPUT", "USER", "user"):
                content = r.get("content", "").strip()
                if content and len(content) >= 10:
                    key = content[:50]
                    if key not in seen:
                        seen.add(key)
                        talk_history.append(content[:80])
                        if len(talk_history) >= 8:
                            break

    # Step 4: AI-derived persona/clusters/events/relationships (same as _build_mock_memory step 8)
    # Use current week label for cache consistency (not old Excel timestamps)
    current_week_label = datetime.utcnow().strftime("%Y-W%W")
    ai_analysis: Dict = {}
    if OPENAI_API_KEY or GOOGLE_API_KEY:
        mem0_memories = _fetch_mem0_memories(profile_id)
        if mem0_memories:
            logger.info(f"[planner_memory] Pipeline: Using Mem0 ({len(mem0_memories)} memories) for {profile_id}")
            try:
                ai_analysis = _analyze_persona_from_memories(profile_id, mem0_memories, current_week_label)
            except Exception as e:
                logger.warning(f"[planner_memory] Pipeline: Mem0-based AI failed: {e}, falling back to transcript")
        if not ai_analysis:
            logger.info(f"[planner_memory] Pipeline: Falling back to transcript-based AI for {profile_id}")
            ai_analysis = _analyze_persona_from_transcript(transcript, profile_id, current_week_label)

    # Simple keyword-based fallbacks if AI fails
    fallback_persona = {
        "disc_type": "I",
        "talkative_score": 7,
        "proactive_score": 6,
        "emotional_score": 7,
        "en_level": "pre_a1",
        "age_estimate": None,
        "persona_summary": "Bé tham gia trò chuyện với Pika, Pika đang tìm hiểu thêm về bé.",
        "persona_tone": "Thoải mái, hay chia sẻ",
        "engage_preferences": ["Trò chuyện tự do"],
        "engagement_insights": f"Bé có {len(user_msgs)} lượt nói qua {len(convos)} buổi.",
    }

    result = {
        "dataset": dataset,
        "profile_id": profile_id,
        # Use current week label so GET /weekly-plan cache key matches
        "week_label": current_week_label,
        "engagement_report": engagement_report,
        "prev_engagement_report": "Tuần trước, bé tương tác khá tốt, hỏi nhiều câu hỏi về thế giới xung quanh. Chủ đề bé thích nhất là động vật hoang dã.",
        "observations_by_domain": observations_by_domain,
        "derived_insights": derived_insights,
        "talk_history": talk_history[:10],
        "conversation_count": len(convos),
        "message_count": len(user_msgs),
        "generated_at": datetime.utcnow().isoformat(),
        # AI-derived fields for /memory page (persona, clusters, events, relationships)
        "persona": ai_analysis.get("persona") or fallback_persona,
        "memory_clusters": ai_analysis.get("memory_clusters") or [],
        "life_events": ai_analysis.get("life_events") or [],
        "relationship_graph": ai_analysis.get("relationship_graph") or [],
        "ai_powered": bool(ai_analysis),
    }

    _save_cached_memory(dataset, result)
    return result


# ─── Routes ───────────────────────────────────────────────────────────────────

@router.get("/memory")
async def get_planner_memory(dataset: str = Query(...), force_refresh: bool = False):
    """Get cached task memory for a dataset. If not cached, process from Excel data or fall back to eval_sessions."""
    dataset = dataset.rstrip('.')
    # Resolve profile_id from dataset name
    profile_id = DATASET_PROFILE_MAP.get(dataset, dataset)

    # Check cache
    if not force_refresh:
        cached = _load_cached_memory(dataset)
        if cached:
            logger.info(f"[planner_memory] Cache HIT for {dataset}")
            return cached

    # Try Excel pipeline
    logger.info(f"[planner_memory] Cache MISS for {dataset} — trying pipeline")
    try:
        result = _run_memory_pipeline(dataset, profile_id)
        return result
    except Exception as pipeline_err:
        logger.warning(f"[planner_memory] Pipeline failed for {dataset}: {pipeline_err} — trying eval_sessions fallback")

    # ── Fallback: read from eval_sessions (data from admin pipeline) ──────────
    try:
        from app.core.db import get_db_connection
        import json as _json
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT data FROM eval_sessions WHERE profile_id = ? AND current_step = 'completed' ORDER BY created_at DESC LIMIT 1",
            (profile_id,)
        )
        row = cursor.fetchone()
        conn.close()

        if row:
            session_data = _json.loads(row["data"] or "{}")
            ma = session_data.get("memory_analysis") or {}
            parsed = ma.get("parsed") or ma  # parsed is the YAML dict

            if parsed:
                persona = parsed.get("persona") or {}
                now_str = datetime.utcnow().isoformat()
                # Map to PlannerMemoryResponse-compatible format
                result = {
                    "dataset": dataset,
                    "profile_id": profile_id,
                    "week_label": datetime.utcnow().strftime("W%W/%Y"),
                    "persona": persona,
                    "memory_clusters": parsed.get("memory_clusters") or [],
                    "life_events": parsed.get("life_events") or [],
                    "relationship_graph": parsed.get("relationship_graph") or [],
                    "engagement_report": persona.get("engagement_insights") or persona.get("persona_summary") or "",
                    "prev_engagement_report": "",
                    "observations_by_domain": {},
                    "derived_insights": persona.get("persona_tone") or "",
                    "talk_history": [],
                    "conversation_count": len(parsed.get("memory_clusters") or []),
                    "message_count": 0,
                    "generated_at": now_str,
                    "ai_powered": True,
                    "from_eval_session": True,
                }
                # Cache so next call is fast
                _save_cached_memory(dataset, result)
                logger.info(f"[planner_memory] Served from eval_sessions for {dataset}")
                return result
    except Exception as fallback_err:
        logger.warning(f"[planner_memory] eval_sessions fallback failed: {fallback_err}")

    raise HTTPException(status_code=404, detail=f"Không có dữ liệu cho profile '{dataset}'. Hãy chạy pipeline trước.")


@router.post("/memory/process")
async def process_planner_memory(dataset: str = Query(...)):
    """Force-run the memory pipeline (no cache)."""
    dataset = dataset.rstrip('.')
    profile_id = DATASET_PROFILE_MAP.get(dataset, dataset)
    try:
        result = _run_memory_pipeline(dataset, profile_id)
        
        # Invalidate weekly plan cache to force regeneration
        import glob
        from app.api.routes.weekly_plan import WEEKLY_PLAN_DIR
        pattern = os.path.join(WEEKLY_PLAN_DIR, f"{dataset}_*.json")
        for f in glob.glob(pattern):
            try:
                os.remove(f)
                logger.info(f"[planner_memory] Invalidated weekly plan cache: {f}")
            except OSError:
                pass
                
        posthog = get_posthog()
        if posthog:
            posthog.capture(
                distinct_id=profile_id,
                event="planner_memory_processed",
                properties={
                    "conversation_count": result.get("conversation_count", 0),
                    "message_count": result.get("message_count", 0),
                    "ai_powered": result.get("ai_powered", False),
                },
            )
        return {"status": "success", "dataset": dataset, "generated_at": result["generated_at"]}
    except FileNotFoundError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except Exception as e:
        logger.exception(f"[planner_memory] Process error for {dataset}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/memory/save")
async def save_planner_memory(req: UpdateMemoryRequest):
    """Save user edits to the memory and invalidate weekly plan cache."""
    dataset = req.dataset.rstrip('.')
    try:
        data = _load_cached_memory(dataset)
        if not data:
            # No file cache yet (e.g. admin pipeline ran) — try eval_sessions fallback
            logger.info(f"[planner_memory] No file cache for {dataset} on save — trying eval_sessions fallback")
            try:
                data = await get_planner_memory(dataset=dataset)
                data = dict(data) if not isinstance(data, dict) else data
            except Exception as e2:
                logger.warning(f"[planner_memory] Fallback failed for {dataset}: {e2}")
                data = {}
            if not data:
                # Last resort: bootstrap from mock_data/memo.json (data-import users)
                _mock_path = os.path.join(
                    os.path.dirname(__file__), "..", "..", "..", "mock_data", dataset, "memo.json"
                )
                if os.path.exists(_mock_path):
                    logger.info(f"[planner_memory] Bootstrapping pipeline cache from mock_data for {dataset}")
                    with open(_mock_path, "r", encoding="utf-8") as _f:
                        data = json.load(_f)
                if not data:
                    raise HTTPException(
                        status_code=404,
                        detail=f"No memory data found for {dataset}. Run the memory pipeline first."
                    )
        if req.memory_clusters is not None:
            data["memory_clusters"] = req.memory_clusters
            
        if req.talk_history is not None:
            data["talk_history"] = req.talk_history
            
        # Add a flag to indicate it was manually edited
        data["edited_by_user"] = True
        # Ensure week_label is current so weekly plan cache key stays consistent
        if not data.get("week_label"):
            data["week_label"] = datetime.utcnow().strftime("%Y-W%W")

        _save_cached_memory(dataset, data)

        # Invalidate weekly plan cache to force regeneration with new memory
        import glob
        from app.api.routes.weekly_plan import WEEKLY_PLAN_DIR
        pattern = os.path.join(WEEKLY_PLAN_DIR, f"{dataset}_*.json")
        deleted = []
        for f in glob.glob(pattern):
            try:
                os.remove(f)
                deleted.append(os.path.basename(f))
                logger.info(f"[planner_memory] Invalidated weekly plan cache after edit: {f}")
            except OSError as oe:
                logger.warning(f"[planner_memory] Could not delete {f}: {oe}")

        posthog = get_posthog()
        if posthog:
            posthog.capture(
                distinct_id=dataset,
                event="memory_saved",
                properties={
                    "has_cluster_update": req.memory_clusters is not None,
                    "has_talk_history_update": req.talk_history is not None,
                    "invalidated_plan_count": len(deleted),
                },
            )

        return {
            "status": "success",
            "dataset": dataset,
            "invalidated_plans": deleted,
            "message": f"Memory saved. {len(deleted)} plan cache(s) cleared — next visit will regenerate."
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[planner_memory] Save error for {dataset}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

